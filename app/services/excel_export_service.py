"""
Service d'export Excel des notes de frais.
Produit un classeur Excel contenant :

- Synthèse
- Notes de frais
- Lignes de frais
- Justificatifs
- Statistiques
"""

from datetime import date

from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from app.models.expense import Expense
from app.models.expense_item import ExpenseItem
from app.models.expense_receipt import ExpenseReceipt
from io import BytesIO
from app.core.config import settings

from openpyxl.utils import get_column_letter

def autofit_columns(sheet):
    
    """
    Adapte automatiquement la largeur des colonnes
    au contenu de la feuille.
    """

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(
            column_cells[0].column
        )
        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )
            except Exception:
                pass
        sheet.column_dimensions[column].width = min(
            max_length + 2,
            80,
        )

def create_excel_export(
    db: Session,
    start_date: date,
    end_date: date,
):

    expenses = (
        db.query(Expense)
        .options(
            joinedload(Expense.user),
            joinedload(Expense.items)
                .joinedload(ExpenseItem.receipts),
        )
        .filter(
            Expense.expense_date >= start_date,
            Expense.expense_date <= end_date,
        )
        .order_by(
            Expense.expense_date,
            Expense.id,
        )
        .all()
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthèse"
    sheet["A1"] = "Exercice"
    sheet["B1"] = (
        f"{start_date:%d/%m/%Y}"
        " au "
        f"{end_date:%d/%m/%Y}"
    )
    sheet["A2"] = "Nombre de notes"
    sheet["B2"] = len(expenses)
    total_items = 0
    total_receipts = 0
    total_amount = 0
    for expense in expenses:
        total_amount += (
            expense.total_amount or 0
        )
        total_items += len(
            expense.items
        )
        for item in expense.items:
            total_receipts += len(
                item.receipts
            )
    sheet["A3"] = "Nombre de lignes"
    sheet["B3"] = total_items
    sheet["A4"] = "Nombre de justificatifs"
    sheet["B4"] = total_receipts
    sheet["A5"] = "Montant total"
    sheet["B5"].number_format = '#,##0.00 €'

    # feuille notes de frais

    sheet = workbook.create_sheet("Notes de frais")
    sheet.append(
        [
            "ID",
            "Date",
            "Utilisateur",
            "Libellé",
            "Statut",
            "Nombre de lignes",
            "Montant",
        ]
    )

    for expense in expenses:
        if not expense.items:
            continue
        note_type = expense.items[0].item_type.name
        if note_type == "KM":
            total_km = sum(
                item.kilometers or 0
                for item in expense.items
            )
            departure = (
                expense.items[0].departure or "?"
            )
            arrival = (
                expense.items[-1].arrival or "?"
            )
            libelle = (
                f"Déplacement : "
                f"{departure} → "
                f"{arrival} "
                f"({total_km:.1f} km)"
            )
        elif note_type == "HOTEL":
            libelle = f"Hôtel : {expense.label}"
        elif note_type == "REPAS":
            libelle = f"Repas : {expense.label}"
        elif note_type == "FOURNITURE":
            libelle = f"Fournitures : {expense.label}"
        elif note_type == "PEAGE":
            libelle = f"Péage : {expense.label}"
        elif note_type == "PARKING":
            libelle = f"Parking : {expense.label}"
        else:
            libelle = expense.label
        sheet.append(
            [
                expense.id,
                expense.expense_date,
                expense.user.full_name,
                libelle,
                expense.status.value,
                len(expense.items),
                float(expense.total_amount or 0),
            ]
        )
        last_row = sheet.max_row
        sheet[f"G{last_row}"].number_format = '#,##0.00 €'

    # feuille justificatifs

    sheet = workbook.create_sheet("Justificatifs")
    sheet.append(
        [
            "NDF",
            "Utilisateur",
            "Fichier",
            "Type MIME",
            "Taille",
            "Numero Archive",
        ]
    )

    for expense in expenses:
        for item in expense.items:
            for receipt in item.receipts:
                sheet.append(
                    [
                        f"NDF-{expense.id:06d}",
                        expense.user.full_name,
                        receipt.original_filename,
                        receipt.content_type,
                        receipt.file_size,
                        (
                            f"{settings.R2_PREFIX}/"
                            f"{settings.R2_EXPENSE_FOLDER}/"
                            f"{expense.expense_date:%Y}/"
                            f"{expense.expense_date:%Y-%m}/"
                            f"NDF-{expense.id:06d}/"
                        ),
                    ]
                )
    
    buffer = BytesIO()
    for sheet in workbook.worksheets:
        autofit_columns(sheet)
    workbook.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    buffer.close()
    return content
