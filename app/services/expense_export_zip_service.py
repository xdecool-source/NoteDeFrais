"""
Service d'export comptable.

Ce service :

- génère un fichier Excel (.xlsx) ;
- génère une archive ZIP contenant :
    - export.xlsx
    - tous les justificatifs.
- ne modifie jamais la base.
"""

import io
import zipfile

from datetime import date

from sqlalchemy.orm import (
    Session,
    joinedload,
)

from app.models.expense import Expense
from app.models.expense_item import ExpenseItem
from app.models.enums import ExpenseStatus

from app.services.receipt_file_service import read_receipt
from app.services.expense_label_service import build_expense_label

from app.core.file_utils import normalize_filename
from app.core.config import settings

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font

def autofit_columns(sheet):
    
    """
    Adapte automatiquement la largeur des colonnes
    au contenu de la feuille.
    """

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(
            column_cells[0].column
        )
        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )
            except Exception:
                pass
        sheet.column_dimensions[column].width = min(
            max_length + 2,
            80,
        )


# categories comptables

CATEGORIES = {
    "KM":
        "Adhérents - Frais de Déplacement / Mission",
    "FOURNITURE":
        "Achats de Fournitures / Boissons",
    "REPAS":
        "Adhérents - Frais de Déplacement / Mission",
    "HOTEL":
        "Adhérents - Frais de Déplacement / Mission",
    "PEAGE":
        "Adhérents - Frais de Déplacement / Mission",
    "PARKING":
        "Adhérents - Frais de Déplacement / Mission",
    "OTHER":
        "Assurances - Divers",
}

# nom des dossiers

DIRECTORIES = {

    "KM": "Deplacement",
    "FOURNITURE": "Fournitures",
    "HOTEL": "Hotel",
    "REPAS": "Repas",
    "OTHER": "Divers",
}

# export

def create_export_zip(
    db: Session,
    start_date: date,
    end_date: date,
):

    # dossier racine du zip

    root_directory = (
        f"Export_{start_date}_{end_date}"
    )

    # recherche des notes

    expenses = (
        db.query(Expense)
        .options(
            joinedload(
                Expense.items
            ).joinedload(
                ExpenseItem.receipts
            ),
            joinedload(
                Expense.user
            ),
        )
        .filter(
            Expense.status == ExpenseStatus.PAID,
            Expense.expense_date >= start_date,
            Expense.expense_date <= end_date,
        )
        .order_by(
            Expense.expense_date,
            Expense.id,
        )
        .all()
    )

    # export comptable :xls
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Export comptable"

    """
    # export detaille csv

    summary_writer.writerow(
        [
            "ID",
            "OPERATION",
            "MONTANT",
            "DATE",
            "LIBELLE",
            "COMPTE_BANCAIRE",
            "CATEGORIE",
            "RAPPROCHE",
            "MODE_REGLEMENT",
            "NUMERO_ARCHIVE",
            "NDF",
            "BENEFICIAIRE",      
        ]
    )
    """
    # export detaille xls

    headers = [
        "ID",
        "OPERATION",
        "MONTANT",
        "DATE",
        "LIBELLE",
        "COMPTE_BANCAIRE",
        "CATEGORIE",
        "RAPPROCHE",
        "MODE_REGLEMENT",
        "NUMERO_ARCHIVE",
        "NDF",
        "BENEFICIAIRE",
    ]

    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    
    # zip

    zip_buffer = io.BytesIO()
    zip_file = zipfile.ZipFile(
        zip_buffer,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    )

    # compteur

    exported_lines = 0

    # parcours des notes
    
    for expense in expenses:

        # nom du dossier de la note

        expense_directory = (
            f"NDF-{expense.id:06d}"
        )

        # ligne comptable
        # Une seule ligne par note de frais

        note_type = expense.items[0].item_type.name
        categorie = CATEGORIES.get(
            note_type,
            "Divers",
        )
        
        # libellé comptable

        libelle = build_expense_label(expense)
        
        # dossier des justificatifs

        has_receipts = any(
            item.receipts
            for item in expense.items
        )
        if has_receipts:
            numero_archive = (
            f"{settings.R2_PREFIX}/"
            f"{settings.R2_EXPENSE_FOLDER}/"
            f"{expense.expense_date:%Y}/"
            f"{expense.expense_date:%Y-%m}/"
            f"NDF-{expense.id:06d}/"
            if has_receipts
            else ""
        )
        else:
            numero_archive = ""
        
        # export comptable
        
        sheet.append(
            [
                "",
                "Dépense",
                f"{expense.total_amount:.2f}",
                expense.expense_date,
                libelle,
                "COMPTE COURANT C.A.",
                categorie,
                "1",
                "virement",
                numero_archive,
                f"NDF-{expense.id:06d}",
                expense.user.username,
            ]
        )
        
        exported_lines += 1
        
                
        # parcours des lignes

        for index, item in enumerate(
            expense.items,
            start=1,
        ):
            directory_name = DIRECTORIES.get(
                item.item_type.name,
                item.item_type.name,
            )
            line_directory = (
                f"{root_directory}/"
                f"{expense_directory}/"
                f"{index:02d}_{directory_name}"
            )

            # justificatifs = numero_archive

            for receipt in item.receipts:
                # nom unique
                receipt_filename = (
                    f"{receipt.id:04d}_"
                    f"{normalize_filename(receipt.original_filename)}"
                )
                archive_filename = (
                    f"{line_directory}/"
                    f"{receipt_filename}"
                )
                # lecture du justificatif
                try:
                    file_data = read_receipt(
                        receipt
                    )    
                except Exception as e:
                    continue
                # ajout au zip
                zip_file.writestr(
                    archive_filename,
                    file_data,
                )
                        
    # ajout du csv dans le zip
    
    
    autofit_columns(sheet)

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"

    for cell in sheet["C"][1:]:
        cell.number_format = '#,##0.00 €'

    for cell in sheet["D"][1:]:
        cell.number_format = "DD/MM/YYYY"

    excel_buffer = io.BytesIO()

    workbook.save(excel_buffer)

    excel_buffer.seek(0)

    zip_file.writestr(
        f"{root_directory}/export.xlsx",
        excel_buffer.getvalue(),
    )

    zip_file.close()

    zip_buffer.seek(0)
    zip_content = zip_buffer.getvalue()

    zip_buffer.close()
    excel_buffer.close()
    
    return (
    zip_content,
    exported_lines,
)
    